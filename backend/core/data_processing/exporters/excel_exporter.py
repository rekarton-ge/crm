"""
Excel экспортер данных.

Этот модуль содержит класс для экспорта данных в формате Excel (XLSX).
"""

import io
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
from django.db.models import Model, QuerySet
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.data_processing.exporters.base import BaseExporter
from core.data_processing.error_handlers import ErrorHandler

# Настройка логгера
logger = logging.getLogger(__name__)


class ExcelExporter(BaseExporter):
    """
    Экспортер данных в формат Excel (XLSX).

    Экспортирует данные из QuerySet или списка моделей в файл Excel
    с возможностью настройки стилей и форматирования.
    """

    format_name = "excel"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    file_extension = "xlsx"

    def __init__(self,
                 fields: Optional[List[str]] = None,
                 exclude_fields: Optional[List[str]] = None,
                 field_labels: Optional[Dict[str, str]] = None,
                 file_name: Optional[str] = None,
                 error_handler: Optional[ErrorHandler] = None,
                 sheet_name: Optional[str] = None,
                 header_style: Optional[Dict[str, Any]] = None,
                 cell_styles: Optional[Dict[str, Dict[str, Any]]] = None,
                 column_widths: Optional[Dict[str, int]] = None,
                 format_values: Optional[Dict[str, callable]] = None,
                 freeze_panes: Optional[str] = None,
                 auto_filter: bool = True):
        """
        Инициализирует Excel экспортер с указанными настройками.

        Args:
            fields: Список полей для экспорта. Если None, экспортируются все поля.
            exclude_fields: Список полей для исключения из экспорта.
            field_labels: Словарь соответствия имен полей и их отображаемых названий.
            file_name: Имя выходного файла без расширения.
            error_handler: Обработчик ошибок для использования.
            sheet_name: Имя листа Excel.
            header_style: Стиль для заголовков.
            cell_styles: Стили для ячеек по имени поля.
            column_widths: Ширина колонок по имени поля.
            format_values: Словарь функций для форматирования значений полей.
            freeze_panes: Ячейка для закрепления областей (например, "A2").
            auto_filter: Применять ли автофильтр к заголовкам.
        """
        super().__init__(fields, exclude_fields, field_labels, file_name, error_handler)

        self.sheet_name = sheet_name or "Данные"

        # Стиль заголовков по умолчанию
        self.header_style = {
            'font': Font(bold=True, size=12, color="FFFFFF"),
            'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }

        # Обновляем стиль заголовков, если передан
        if header_style:
            self.header_style.update(header_style)

        self.cell_styles = cell_styles or {}
        self.column_widths = column_widths or {}
        self.format_values = format_values or {}
        self.freeze_panes = freeze_panes
        self.auto_filter = auto_filter

    def format_value(self, value: Any, field_name: str) -> Any:
        """
        Форматирует значение поля для экспорта в Excel.

        Args:
            value: Значение поля
            field_name: Имя поля

        Returns:
            Any: Отформатированное значение поля
        """
        # Если есть функция форматирования для данного поля, применяем ее
        if field_name in self.format_values and callable(self.format_values[field_name]):
            return self.format_values[field_name](value)

        # None остается None для Excel
        if value is None:
            return ""

        # Даты и время оставляем как есть для правильного форматирования в Excel
        if isinstance(value, (date, datetime)):
            return value

        # Преобразуем в строку, если значение не является базовым типом Excel
        if not isinstance(value, (str, int, float, bool, date, datetime)):
            return str(value)

        return value

    def get_row(self, obj: Any) -> List[Any]:
        """
        Получает список значений объекта для экспорта в Excel.

        Args:
            obj: Объект для экспорта

        Returns:
            List[Any]: Список значений полей объекта
        """
        return [
            self.format_value(self.get_value(obj, field['name']), field['name'])
            for field in self.export_fields
        ]

    def apply_styles(self, worksheet: Worksheet, headers: List[str]) -> None:
        """
        Применяет стили к листу Excel.

        Args:
            worksheet: Лист Excel
            headers: Список заголовков
        """
        # Применяем стиль к заголовкам
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx)

            # Применяем стили заголовка
            for style_attr, style_value in self.header_style.items():
                setattr(cell, style_attr, style_value)

        # Устанавливаем ширину колонок
        for col_idx, field in enumerate(self.export_fields, 1):
            column_letter = get_column_letter(col_idx)
            field_name = field['name']

            # Устанавливаем ширину колонки, если задана
            if field_name in self.column_widths:
                worksheet.column_dimensions[column_letter].width = self.column_widths[field_name]
            else:
                # Автоматическая ширина на основе длины заголовка (минимум 10)
                worksheet.column_dimensions[column_letter].width = max(len(field['verbose_name']) + 2, 10)

        # Закрепление областей
        if self.freeze_panes:
            worksheet.freeze_panes = self.freeze_panes
        else:
            # По умолчанию закрепляем строку заголовка
            worksheet.freeze_panes = "A2"

        # Автофильтр
        if self.auto_filter:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{worksheet.max_row}"

    def create_workbook(self, headers: List[str], data: List[List[Any]]) -> Workbook:
        """
        Создает книгу Excel с данными.

        Args:
            headers: Список заголовков
            data: Список строк данных

        Returns:
            Workbook: Созданная книга Excel
        """
        # Создаем новую книгу Excel
        workbook = Workbook()

        # Берем активный лист
        worksheet = workbook.active

        # Устанавливаем имя листа
        worksheet.title = self.sheet_name

        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)

        # Записываем данные
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

                # Применяем стили к ячейкам, если заданы
                field_name = self.export_fields[col_idx - 1]['name']
                if field_name in self.cell_styles:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_style = self.cell_styles[field_name]

                    for style_attr, style_value in cell_style.items():
                        setattr(cell, style_attr, style_value)

        # Применяем стили
        self.apply_styles(worksheet, headers)

        return workbook

    def export_data(self, queryset: Union[QuerySet, List[Model]]) -> bytes:
        """
        Экспортирует данные в формат Excel (XLSX).

        Args:
            queryset: QuerySet или список моделей для экспорта

        Returns:
            bytes: Байтовое представление Excel-файла
        """
        # Подготовка данных для экспорта
        headers, data = self.prepare_data(queryset)

        # Создаем книгу Excel
        workbook = self.create_workbook(headers, data)

        # Сохраняем книгу в буфер
        output = io.BytesIO()
        workbook.save(output)

        # Возвращаем данные как байты
        return output.getvalue()

    @classmethod
    def export_queryset(cls,
                        queryset: Union[QuerySet, List[Model]],
                        fields: Optional[List[str]] = None,
                        file_name: Optional[str] = None,
                        **kwargs) -> bytes:
        """
        Статический метод для быстрого экспорта QuerySet в Excel.

        Args:
            queryset: QuerySet или список моделей для экспорта
            fields: Список полей для экспорта
            file_name: Имя выходного файла
            **kwargs: Дополнительные аргументы для конструктора ExcelExporter

        Returns:
            bytes: Байтовое представление Excel-файла
        """
        exporter = cls(fields=fields, file_name=file_name, **kwargs)
        return exporter.export_data(queryset)