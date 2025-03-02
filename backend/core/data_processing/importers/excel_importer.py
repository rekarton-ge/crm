"""
Excel импортер данных.

Этот модуль содержит класс для импорта данных из Excel-файлов в модели Django.
"""

import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Type, Union

import openpyxl
from django.db.models import Model
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.data_processing.error_handlers import (
    ErrorCategory,
    ErrorHandler,
    ErrorSeverity,
    ProcessingError,
    ProcessingResult
)
from core.data_processing.importers.base import BaseImporter

# Настройка логгера
logger = logging.getLogger(__name__)


class ExcelImporter(BaseImporter):
    """
    Импортер данных из Excel-файлов.

    Позволяет импортировать данные из Excel-файлов (XLSX, XLS)
    в модели Django с гибкой настройкой параметров импорта.
    """

    format_name = "excel"
    supported_extensions = ["xlsx", "xls"]

    def __init__(self,
                 model_class: Type[Model],
                 mapping: Optional[Dict[str, str]] = None,
                 sheet_name: Optional[str] = None,
                 sheet_index: int = 0,
                 has_header: bool = True,
                 detect_header: bool = True,
                 header_row: int = 1,
                 data_start_row: Optional[int] = None,
                 data_end_row: Optional[int] = None,
                 ignore_empty_rows: bool = True,
                 **kwargs):
        """
        Инициализирует Excel импортер с указанными настройками.

        Args:
            model_class: Класс модели Django для импорта данных.
            mapping: Словарь соответствия полей в файле полям модели.
            sheet_name: Имя листа Excel для импорта.
            sheet_index: Индекс листа Excel (используется, если sheet_name не указан).
            has_header: Имеет ли файл строку заголовка.
            detect_header: Пытаться ли автоматически определить заголовки.
            header_row: Номер строки заголовка (если has_header=True, нумерация с 1).
            data_start_row: Номер строки начала данных (если не указан, равен header_row + 1).
            data_end_row: Номер строки окончания данных (если не указан, до конца листа).
            ignore_empty_rows: Игнорировать ли пустые строки.
            **kwargs: Дополнительные аргументы для базового импортера.
        """
        super().__init__(model_class, mapping, **kwargs)

        self.sheet_name = sheet_name
        self.sheet_index = sheet_index
        self.has_header = has_header
        self.detect_header = detect_header
        self.header_row = header_row
        self.data_start_row = data_start_row or (header_row + 1 if has_header else 1)
        self.data_end_row = data_end_row
        self.ignore_empty_rows = ignore_empty_rows

        # Для хранения заголовков Excel
        self.headers = []
        self.column_indices = {}

    def get_sheet(self, workbook: Workbook) -> Worksheet:
        """
        Получает лист Excel для обработки.

        Args:
            workbook: Книга Excel.

        Returns:
            Worksheet: Лист Excel.

        Raises:
            ValueError: Если указанный лист не найден.
        """
        # Если указано имя листа, ищем его
        if self.sheet_name:
            if self.sheet_name in workbook.sheetnames:
                return workbook[self.sheet_name]
            else:
                avail_sheets = ", ".join(workbook.sheetnames)
                raise ValueError(f"Лист '{self.sheet_name}' не найден. Доступные листы: {avail_sheets}")

        # Иначе используем индекс листа
        if self.sheet_index < len(workbook.sheetnames):
            return workbook.worksheets[self.sheet_index]
        else:
            raise ValueError(
                f"Индекс листа {self.sheet_index} вне диапазона (всего листов: {len(workbook.sheetnames)})")

    def get_cell_value(self, cell: Cell) -> Any:
        """
        Получает значение ячейки Excel.

        Args:
            cell: Ячейка Excel.

        Returns:
            Any: Значение ячейки.
        """
        # Если ячейка None или пустая
        if cell is None:
            return None

        value = cell.value

        # Если ячейка пустая
        if value is None or value == "":
            return None

        # Преобразование типов данных
        if isinstance(value, datetime):
            # Если ячейка имеет только дату, возвращаем только дату
            if value.hour == 0 and value.minute == 0 and value.second == 0:
                return value.date()

        return value

    def is_row_empty(self, row_data: Dict[str, Any]) -> bool:
        """
        Проверяет, является ли строка пустой.

        Args:
            row_data: Словарь данных строки.

        Returns:
            bool: True, если строка пустая, иначе False.
        """
        # Строка считается пустой, если все значения None или пустые строки
        return all(value is None or value == "" for value in row_data.values())

    def auto_detect_header(self, worksheet: Worksheet) -> List[str]:
        """
        Автоматически определяет заголовки из строки Excel.

        Args:
            worksheet: Лист Excel.

        Returns:
            List[str]: Список заголовков.
        """
        headers = []
        header_cells = list(worksheet.rows)[self.header_row - 1]

        for cell in header_cells:
            value = self.get_cell_value(cell)
            if value is not None:
                headers.append(str(value).strip())
            else:
                # Если в ячейке заголовка нет значения, используем индекс столбца
                headers.append(f"Column_{cell.column}")

        return headers

    def create_mapping_from_headers(self, headers: List[str]) -> Dict[str, str]:
        """
        Создает отображение полей на основе заголовков Excel.

        Args:
            headers: Список заголовков Excel.

        Returns:
            Dict[str, str]: Словарь отображения полей.
        """
        mapping = {}

        # Получаем список полей модели в нижнем регистре
        model_fields_lower = [field.lower() for field in self.model_fields]

        for col_idx, header in enumerate(headers):
            # Нормализуем заголовок
            normalized_header = str(header).lower().strip()

            # Проверяем точное совпадение заголовка с полем модели
            if normalized_header in self.model_fields:
                mapping[col_idx] = normalized_header
                continue

            # Проверяем, совпадает ли заголовок с полем модели без учета регистра
            if normalized_header in model_fields_lower:
                idx = model_fields_lower.index(normalized_header)
                mapping[col_idx] = self.model_fields[idx]
                continue

            # Проверяем, содержит ли заголовок имя поля модели
            for field in self.model_fields:
                if field.lower() in normalized_header.lower():
                    mapping[col_idx] = field
                    break

        return mapping

    def read_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Читает данные из Excel-файла.

        Args:
            file_path: Путь к Excel-файлу.

        Returns:
            Dict[str, Any]: Словарь с информацией о файле Excel.
        """
        try:
            # Открываем книгу Excel с данными
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Получаем лист
            worksheet = self.get_sheet(workbook)

            return {
                'workbook': workbook,
                'worksheet': worksheet
            }

        except Exception as e:
            logger.error(f"Ошибка при чтении Excel-файла: {str(e)}")
            raise

    def process_excel_data(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Обрабатывает данные из листа Excel и преобразует их в список словарей.

        Args:
            worksheet: Лист Excel.

        Returns:
            List[Dict[str, Any]]: Список словарей с данными.
        """
        result = []

        # Получаем размеры листа
        max_row = worksheet.max_row

        # Если задан data_end_row, ограничиваем максимальную строку
        if self.data_end_row and self.data_end_row < max_row:
            max_row = self.data_end_row

        # Определяем заголовки, если нужно
        if self.has_header:
            if self.header_row > max_row:
                raise ValueError(
                    f"Номер строки заголовка ({self.header_row}) превышает количество строк в листе ({max_row})")

            self.headers = self.auto_detect_header(worksheet)

            # Создаем отображение индексов столбцов для заголовков
            self.column_indices = {header: i for i, header in enumerate(self.headers)}

            # Если mapping не задан и нужно определить его автоматически
            if not self.mapping and self.detect_header:
                self.mapping = self.create_mapping_from_headers(self.headers)
                logger.info(f"Автоматически определено отображение полей: {self.mapping}")

        # Обрабатываем строки данных
        for row_idx in range(self.data_start_row, max_row + 1):
            # Получаем ячейки строки
            row_cells = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]

            row_data = {}

            # Если есть заголовки и нет явного отображения
            if self.headers and not self.mapping:
                for col_idx, header in enumerate(self.headers):
                    if col_idx < len(row_cells):
                        cell = row_cells[col_idx]
                        row_data[header] = self.get_cell_value(cell)
            # Если есть отображение
            elif self.mapping:
                # Если отображение по индексам
                if all(isinstance(k, int) for k in self.mapping.keys()):
                    for col_idx, model_field in self.mapping.items():
                        if col_idx < len(row_cells):
                            cell = row_cells[col_idx]
                            row_data[col_idx] = self.get_cell_value(cell)
                # Если отображение по именам колонок
                elif self.headers:
                    for header, model_field in self.mapping.items():
                        if header in self.column_indices:
                            col_idx = self.column_indices[header]
                            if col_idx < len(row_cells):
                                cell = row_cells[col_idx]
                                row_data[header] = self.get_cell_value(cell)
            # Если нет ни заголовков, ни отображения
            else:
                for col_idx, cell in enumerate(row_cells):
                    row_data[col_idx] = self.get_cell_value(cell)

            # Пропускаем пустые строки, если нужно
            if self.ignore_empty_rows and self.is_row_empty(row_data):
                continue

            result.append(row_data)

        return result

    def read_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Читает данные из Excel-файла и преобразует их в список словарей.

        Args:
            file_path: Путь к Excel-файлу.

        Returns:
            List[Dict[str, Any]]: Список словарей с данными.
        """
        # Читаем Excel-файл
        excel_data = self.read_excel(file_path)

        # Обрабатываем данные Excel
        return self.process_excel_data(excel_data['worksheet'])

    @classmethod
    def import_from_excel(cls,
                          file_path: str,
                          model_class: Type[Model],
                          mapping: Optional[Dict[str, str]] = None,
                          **kwargs) -> ProcessingResult:
        """
        Статический метод для быстрого импорта из Excel-файла.

        Args:
            file_path: Путь к Excel-файлу.
            model_class: Класс модели Django для импорта данных.
            mapping: Словарь соответствия полей в файле полям модели.
            **kwargs: Дополнительные аргументы для конструктора ExcelImporter.

        Returns:
            ProcessingResult: Результат импорта.
        """
        importer = cls(model_class=model_class, mapping=mapping, **kwargs)
        return importer.import_file(file_path)